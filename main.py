import os
import re
import io
import sys
import requests
import pandas as pd
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import Tuple
from win32com.client import Dispatch

DAY_TH_TO_CODE = {
    "จันทร์": "M", "อังคาร": "T",
    "พุธ": "W", "พฤหัสบดี": "TH",
    "ศุกร์": "F"
}

def to_csv_url(raw_url: str, gid: str) -> str:
    m = re.search(r"/d/([\-\w]+)", raw_url)
    if not m or not gid.isdigit():
        raise ValueError("Invalid URL or GID")
    return f"https://docs.google.com/spreadsheets/d/{m.group(1)}/export?format=csv&gid={gid}"

def load_priority_mapping(csv_url: str) -> dict[str, int]:
    resp = requests.get(csv_url, timeout=10)
    resp.raise_for_status()

    try:
        df = pd.read_csv(io.BytesIO(resp.content), encoding="cp874", header=None)
    except:
        df = pd.read_csv(io.BytesIO(resp.content), encoding="utf-8-sig", header=None)

    df.columns = ["num", "letter"]

    mapping = {}
    for _, row in df.iterrows():
        try:
            key = str(row["letter"]).strip().upper()
            val = int(row["num"])
            mapping[key] = val
        except Exception:
            continue

    print("\n=== ลำดับตึกจากแผ่นที่ 2 ===")
    for k, v in sorted(mapping.items()):
        print(f"ตัวอักษร {k} = {v}")

    return mapping





def room_key_full(room: str):
    return (room.strip(),)

def room_sort_key(room: str):
    parts = re.findall(r'\d+', room)
    return tuple(map(int, parts)) if parts else (9999,)

def short_room(room: str) -> str:
    parts = room.split("/")
    return parts[0] + parts[-1]

def expand_rooms(raw: str) -> list[str]:
    raw = raw.replace("–", "-").replace("—", "-").strip()
    if not raw:
        return []

    if "/" not in raw:
        return [raw]

    m = re.match(r"([^/]+)/(\d+)/(.+)", raw)
    if m:
        prefix = f"{m.group(1).strip()}/{m.group(2).strip()}"
        nums = m.group(3).strip()
    else:
        m2 = re.match(r"(.+?)/(.+)", raw)
        if not m2:
            return []
        prefix, nums = m2.group(1).strip(), m2.group(2).strip()

    out = []
    for part in re.split(r"\s*,\s*", nums):
        if "-" in part:
            try:
                a, b = map(int, part.split("-", 1))
                out.extend([f"{prefix}/{i}" for i in range(a, b + 1)])
            except ValueError:
                continue
        elif part.isdigit():
            out.append(f"{prefix}/{part}")
    return out

def load_subjects(csv_url: str) -> list[dict]:
    resp = requests.get(csv_url, timeout=10)
    resp.raise_for_status()
    try:
        df = pd.read_csv(io.BytesIO(resp.content), encoding="cp874")
    except:
        df = pd.read_csv(io.BytesIO(resp.content), encoding="utf-8-sig")
    df.columns = df.columns.str.strip()

    def find_col(keys):
        for c in df.columns:
            if any(k.lower() in c.lower() for k in keys):
                return c
        return None

    code_c = find_col(["รหัสวิชา", "code"])
    cred_c = find_col(["หน่วยกิต", "credit"])
    teacher_c = find_col(["ครู", "teacher"])
    weight_c = find_col(["น้ำหนัก", "weight"])
    room_std_c = find_col(["ห้องนักเรียน"])
    sum_idx = df.columns.get_loc(find_col(["สรุปห้อง"]))

    subjects = []
    df = df[df[code_c].notna()]
    df[cred_c] = df[cred_c].fillna(0).astype(float)
    df[weight_c] = df[weight_c].fillna(0).astype(float)

    for _, row in df.iterrows():
        code = str(row[code_c]).strip()
        credit = float(row[cred_c])
        teacher = str(row[teacher_c]).strip()
        weight = float(row[weight_c])

        raw_group = str(row[room_std_c]).strip()
        group_list = expand_rooms(raw_group) if raw_group else []
        if not group_list:
            continue

        actual_room_list = []
        for i in range(sum_idx, len(row)):
            val = str(row.iloc[i]).strip()
            if val and val.lower() not in ("-", "nan"):
                for part in re.split(r"[;\n]+", val):
                    actual_room_list += expand_rooms(part.strip())
        actual_room_list = sorted(set(actual_room_list), key=room_key_full)

        print(f"[INFO] วิชา {code} ใช้ห้องจริง: {', '.join(actual_room_list) if actual_room_list else 'ไม่มี'}")

        for gr in group_list:
            subjects.append({
                "code": code,
                "credit": credit,
                "teacher": teacher,
                "weight": weight,
                "group": gr,
                "actual_room": actual_room_list
            })


    return subjects

def build_tasks(subs: list[dict]) -> list[dict]:
    return sorted(
        [s.copy() for s in subs for _ in range(max(1, int(s["credit"] * 2)))],
        key=lambda x: (-x["weight"], *room_key_full(x["group"]), x["code"])
    )



def find_actual_room(subject: dict, group_short: str) -> str:
    if not subject.get("actual_room"):
        return group_short
    for r in subject["actual_room"]:
        if r == subject["group"]:
            return r
    for r in subject["actual_room"]:
        if short_room(r) == group_short:
            return r
    return group_short

def convert_room_letter_to_number(room: str, mapping: dict[str, int]) -> str:
    m = re.match(r"([A-Z]+)(\d+)", room.strip().upper())
    if not m:
        return room
    letter, number = m.groups()
    new_prefix = str(mapping.get(letter, letter))
    return f"{new_prefix}{number}"



def schedule_room(group: str,
                  subjects: list[dict],
                  slots_used: dict[str, set],
                  teacher_slots: dict[str, dict],
                  locks: list[dict],
                  room_priority_map: dict[str, int]) -> Tuple[pd.DataFrame, list[dict]]:

    days = ["M", "T", "W", "TH", "F"]
    day_map = dict(zip(days, ["Mon", "Tue", "Wed", "Thu", "Fri"]))
    table = pd.DataFrame("", index=day_map.values(), columns=range(1, 12), dtype=str)

    def fmt(tr_rooms, room_conf, total_tasks):
        if total_tasks == 0:
            return "ไม่มีวิชา"
        parts = []
        if tr_rooms:
            parts.append("ครูชนที่ " + ",".join(sorted({short_room(r) for r in tr_rooms})))
        if room_conf:
            parts.append("ห้องเต็ม")
        return "ไม่ได้(" + ", ".join(parts) + ")" if parts else "ไม่ได้"

    for lk in locks:
        if "ALL" in lk["rooms"] or group in lk["rooms"]:
            d = DAY_TH_TO_CODE[lk["day"]]
            p = int(lk["period"])
            sl = f"{d}{p}"
            table.at[day_map[d], p] = lk["name"]
            slots_used[sl].add(group)

    subs = [s for s in subjects if s["group"] == group]
    total_credit = sum(s["credit"] for s in subs)
    red_n = 15 if total_credit <= 18 else 20
    tasks = []
    for s in subs:
        tasks += [s.copy() for _ in range(max(1, int(s["credit"] * 2)))]

    def sort_by_building(task_list):
        return sorted(task_list, key=lambda x: (parse_room_building(x["group"]), -x["weight"], x["code"]))

    red, orange, yellow = tasks[:red_n], tasks[red_n:red_n + 10], tasks[red_n + 10:]
    all_tasks = sort_by_building(red) + sort_by_building(orange) + sort_by_building(yellow)
    ordered_slots = generate_schedule_slots(total_credit)

    for sl in ordered_slots:
        d, p = re.match(r"([A-Z]+)(\d+)", sl).groups()
        p = int(p)
        if table.at[day_map[d], p] != "":
            continue

        tr_rooms, room_conf, placed = set(), False, False

        for t in list(all_tasks):
            if sl in teacher_slots.get(t["teacher"], {}):
                tr_rooms.add(teacher_slots[t["teacher"]][sl])
                continue

            for actual in t.get("actual_room", []):
                if actual in slots_used[sl]:
                    room_conf = True
                    continue

                # ✅ จัดได้ → ใส่ตารางแล้ว break
                actual_room_converted = convert_room_letter_to_number(actual, room_priority_map)
                table.at[day_map[d], p] = {
                    "code": t['code'],
                    "teacher": t['teacher'],
                    "room": actual_room_converted
                }

                slots_used[sl].update({t["teacher"], actual})
                teacher_slots.setdefault(t["teacher"], {})[sl] = actual
                all_tasks.remove(t)
                placed = True
                break  # ✅ break ออกจาก actual_room → แต่ยังอยู่ในลูป task

            if placed:
                break  # ✅ break ตรงนี้เท่านั้น เมื่อจัดได้แล้ว

        if not placed:
            table.at[day_map[d], p] = fmt(tr_rooms, room_conf, len(all_tasks))

    return table, all_tasks


def generate_schedule_slots(total_credit: float) -> list[str]:
    days = ["M", "T", "W", "TH", "F"]
    order = []

    def append_red():
        for d in days:
            order.append(f"{d}2")
            order.append(f"{d}3")

    def append_red_extra():
        for d in days:
            order.append(f"{d}6")
            order.append(f"{d}7")

    def append_orange():
        for p in [4, 1]:
            for d in days:
                order.append(f"{d}{p}")

    def append_yellow():
        for p in [8, 9, 10, 11]:
            for d in days:
                order.append(f"{d}{p}")

    if total_credit <= 18:
        append_red()          # M2 M3 T2 T3 ...
        for d in days:
            order.append(f"{d}6")   # M6 T6 ...
        for d in days:
            order.append(f"{d}7")   # M7 T7 ... orange
        append_orange()      # 4, 1
        append_yellow()
    else:
        append_red()         # M2 M3 T2 T3 ...
        append_red_extra()   # M6 M7 ...
        append_orange()      # 4, 1
        append_yellow()

    return order



def interleave(lst: list) -> list:
    half = len(lst) // 2
    result = []
    for i in range(half):
        result.append(lst[i])
        if i + half < len(lst):
            result.append(lst[i + half])
    if len(lst) % 2:
        result.append(lst[-1])
    return result

def room_distance(room1: str, room2: str) -> int:
    return 0 if room1 == room2 else 1

def group_by_teacher_tasks(tasks: list[dict]) -> dict[str, list[dict]]:
    groups = {}
    for t in tasks:
        groups.setdefault(t["teacher"], []).append(t)
    return groups

def find_best_room_for_teacher(tasks: list[dict], teacher: str, sl: str, teacher_slots: dict[str, dict]) -> dict | None:
    prev_rooms = teacher_slots.get(teacher, {}).values()
    best = None
    best_dist = float("inf")

    for t in tasks:
        if t["teacher"] != teacher:
            continue
        r = t["group"]
        dist = min((room_distance(r, pr) for pr in prev_rooms), default=0)
        if dist < best_dist:
            best = t
            best_dist = dist

    return best

def parse_room_building(room: str) -> str:
    return room.strip().split("/")[0].upper()


class SchedulerApp:
    def __init__(self, master):
        self.master = master
        master.title("Timetable Scheduler")
        self.locks = []
        self.labels = {}
        self.subjects = []
        self.rooms = []
        self.slots_used = {}
        self.teacher_slots = {}
        self.lock_name = tk.StringVar()
        self.lock_rooms = tk.StringVar()
        self.lock_day = tk.StringVar()
        self.lock_period = tk.StringVar()
        self.url_var = tk.StringVar()
        self.gid_var = tk.StringVar()
        self.gid_priority_var = tk.StringVar()
        self.priority_mapping = {}
        self.room_priority_map = {}

        # GUI layout
        frm = ttk.Frame(master, padding=10)
        frm.pack()

        ttk.Label(frm, text="Google Sheet URL:").grid(row=0, column=0)
        ttk.Entry(frm, textvariable=self.url_var, width=40).grid(row=0, column=1)
        ttk.Label(frm, text="GID:").grid(row=0, column=2)
        ttk.Entry(frm, textvariable=self.gid_var, width=8).grid(row=0, column=3)
        ttk.Button(frm, text="โหลดข้อมูล", command=self.load_data).grid(row=0, column=4, padx=5)

        ttk.Label(frm, text="GID ลำดับตึก:").grid(row=0, column=5)
        ttk.Entry(frm, textvariable=self.gid_priority_var, width=8).grid(row=0, column=6)

        ttk.Label(frm, text="ชื่อคาบ:").grid(row=1, column=0)
        ttk.Entry(frm, textvariable=self.lock_name).grid(row=1, column=1)
        ttk.Label(frm, text="ห้อง:").grid(row=1, column=2)
        ttk.Entry(frm, textvariable=self.lock_rooms).grid(row=1, column=3)

        ttk.Label(frm, text="ตัวอย่างการกรอกห้อง: ม.1/1-11 (ห้อง ม.1/1 ถึง ม.1/11) หรือพิมพ์ * เพื่อหมายถึงทุกห้อง", foreground="blue", font=("Arial", 9)).grid(row=1,column=4,columnspan=3,sticky="w",padx=5)


        ttk.Label(frm, text="วัน:").grid(row=2, column=0)
        ttk.Combobox(frm, textvariable=self.lock_day, values=list(DAY_TH_TO_CODE.keys()), state="readonly").grid(row=2, column=1)
        ttk.Label(frm, text="คาบ:").grid(row=2, column=2)
        ttk.Entry(frm, textvariable=self.lock_period).grid(row=2, column=3)
        ttk.Button(frm, text="เพิ่ม Lock", command=self.add_lock).grid(row=2, column=4)
        ttk.Button(frm, text="Export PDF", command=self.export_rooms_pdf).grid(row=3, column=3)

        ttk.Label(frm, text="กลุ่มเรียน:").grid(row=3, column=0)
        self.cb = ttk.Combobox(frm, state="readonly")
        self.cb.grid(row=3, column=1)
        self.cb.bind("<<ComboboxSelected>>", lambda e: self.update_grid())
        ttk.Button(frm, text="Export Excel", command=self.export_rooms_excel).grid(row=3, column=2)

        # ตาราง
        days_en = ["Mon", "Tue", "Wed", "Thu", "Fri"]
        grid = ttk.Frame(master)
        grid.pack()
        tk.Label(grid, text="Day/Period", width=12, relief="solid").grid(row=0, column=0)
        for p in range(1, 12):
            tk.Label(grid, text=str(p), width=8, relief="solid").grid(row=0, column=p)
        for i, d in enumerate(days_en, 1):
            tk.Label(grid, text=d, width=12, relief="solid").grid(row=i, column=0)
            for p in range(1, 12):
                lbl = tk.Label(grid, text="", width=20, height=4, relief="solid", anchor="w", justify="left",font=("Arial", 10))
                lbl.grid(row=i, column=p)
                self.labels[(d, p)] = lbl


    def add_lock(self):
        name = self.lock_name.get().strip()
        rooms = self.lock_rooms.get().strip()
        day = self.lock_day.get().strip()
        pr_raw = self.lock_period.get().strip()
        if not (name and rooms and day and pr_raw):
            return
        room_list = ["ALL"] if rooms == "*" else expand_rooms(rooms)
        periods = []
        for part in re.split(r"\s*,\s*", pr_raw):
            if "-" in part:
                a, b = map(int, part.split("-", 1))
                periods += list(range(a, b + 1))
            else:
                periods.append(int(part))
        for p in periods:
            self.locks.append({
                "name": name,
                "rooms": room_list,
                "day": day,
                "period": str(p)
            })

    def load_data(self):
        url, gid = self.url_var.get().strip(), self.gid_var.get().strip()
        try:
            self.subjects = load_subjects(to_csv_url(url, gid))
            group_subjects = {}
            for s in self.subjects:
                gr = s["group"]
                subj = f"{s['code']} ({s['credit']} หน่วยกิต) ครู: {s['teacher']} ห้อง: {', '.join(s['actual_room']) if s['actual_room'] else 'ไม่มี'}"
                group_subjects.setdefault(gr, []).append(subj)
            self.rooms = sorted({s["group"] for s in self.subjects},key=room_sort_key)  # ✅ เปลี่ยน key เป็น room_sort_key
            self.cb["values"] = self.rooms
            if self.rooms:
                self.cb.current(0)
                self.update_grid()

                priority_gid = self.gid_priority_var.get().strip()
                if priority_gid:
                    try:
                        priority_url = to_csv_url(url, priority_gid)
                        self.priority_mapping = load_priority_mapping(priority_url)
                    except Exception as e:
                        print(f"[ERROR] โหลดลำดับตึกไม่สำเร็จ: {e}")

                print("\n=== รายวิชาแยกตามกลุ่มเรียน ===")

                for gr in sorted(self.rooms, key=room_sort_key):
                    subs_sorted = sorted(
                        [s for s in self.subjects if s["group"] == gr],
                        key=lambda x: -x["weight"]
                    )
                    total_credits = sum(s["credit"] for s in subs_sorted)
                    print(f"\n[กลุ่ม {gr}] รวม {total_credits} หน่วยกิต:")
                    for s in subs_sorted:
                        actual = ", ".join(s["actual_room"]) if s["actual_room"] else "-"
                        print(
                            f" - {s['code']} | ครู: {s['teacher']} | ห้อง: {actual} | หน่วยกิต: {s['credit']} | น้ำหนัก: {s['weight']}")

            messagebox.showinfo("Loaded", f"Loaded {len(self.subjects)} entries → {len(self.rooms)} rooms.")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def update_grid(self):
        group = self.cb.get()
        if not group:
            return

        self.slots_used = {f"{d}{p}": set() for d in DAY_TH_TO_CODE.values() for p in range(1, 12)}
        self.teacher_slots = {}
        df = None

        # ✅ วนทุกกลุ่ม → หา df ของกลุ่มที่เลือกเท่านั้น
        for gr in self.rooms:
            df_temp, _ = schedule_room(gr, self.subjects, self.slots_used, self.teacher_slots, self.locks,
                                       self.priority_mapping)
            if gr == group:
                df = df_temp

        # ✅ แสดงผลกลุ่มเดียว
        for (d, p), lbl in self.labels.items():
            val = df.at[d, p]
            if isinstance(val, str):
                lbl["text"] = val.strip()
            elif isinstance(val, dict):
                room_conv = convert_room_letter_to_number(val['room'], self.priority_mapping)
                teacher_name = val['teacher'].split()[0]  # ⬅️ ตัดเฉพาะชื่อหน้า
                lbl["text"] = f"รหัส: {val['code']}\nครู: {teacher_name}\nห้อง: {room_conv}"

            else:
                lbl["text"] = ""

    def export_rooms_excel(self):
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
        if not path:
            return
        self.export_rooms_to(path)
        messagebox.showinfo("Exported", f"Saved Excel → {path}")

    def export_rooms_pdf(self):
        path = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF", "*.pdf")])
        if not path:
            return

        # ▶ แสดง popup สถานะ
        loading_popup = tk.Toplevel(self.master)
        loading_popup.title("กำลังสร้าง PDF...")
        loading_popup.resizable(False, False)

        # ▶ สร้าง Label
        tk.Label(loading_popup, text="กำลังสร้างไฟล์ PDF กรุณารอสักครู่...", padx=20, pady=20).pack()

        # ▶ อัปเดตหน้าต่างก่อนหาตำแหน่ง
        loading_popup.update_idletasks()

        # ▶ คำนวณตำแหน่งกลางจอ
        screen_width = loading_popup.winfo_screenwidth()
        screen_height = loading_popup.winfo_screenheight()
        win_width = loading_popup.winfo_width()
        win_height = loading_popup.winfo_height()

        x = (screen_width // 2) - (win_width // 2)
        y = (screen_height // 2) - (win_height // 2)
        loading_popup.geometry(f"+{x}+{y}")

        loading_popup.update()

        try:
            import tempfile
            from openpyxl import load_workbook
            from openpyxl.worksheet.page import PageMargins
            from win32com.client import Dispatch

            tmp_xlsx = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False).name
            self.export_rooms_to(tmp_xlsx)

            excel = Dispatch("Excel.Application")
            excel.Visible = False
            wb = excel.Workbooks.Open(tmp_xlsx)

            for i, sheet in enumerate(wb.Sheets):
                sheet.PageSetup.Orientation = 2  # xlLandscape
                sheet.PageSetup.Zoom = False
                sheet.PageSetup.FitToPagesWide = 1
                sheet.PageSetup.FitToPagesTall = 1

            wb.ExportAsFixedFormat(0, path)  # 0 = PDF
            wb.Close(False)
            excel.Quit()
            messagebox.showinfo("Exported", f"Saved PDF → {path}")

        except Exception as e:
            messagebox.showerror("Error", str(e))
        finally:
            loading_popup.destroy()




    def export_rooms_to(self, path: str):
        wb = Workbook()
        days_en = ["Mon", "Tue", "Wed", "Thu", "Fri"]
        day_en_to_th = {
            "Mon": "จันทร์", "Tue": "อังคาร", "Wed": "พุธ", "Thu": "พฤหัสบดี", "Fri": "ศุกร์"
        }

        thin = Side(border_style="thin", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        slots_used = {f"{d}{p}": set() for d in DAY_TH_TO_CODE.values() for p in range(1, 12)}
        teacher_slots = {}

        for idx, group in enumerate(sorted(self.rooms, key=room_sort_key)):
            ws = wb.active if idx == 0 else wb.create_sheet()
            ws.title = group.replace("/", "-")[:31]

            ws.cell(1, 1, f"ห้อง: {group}").border = border
            ws.cell(1, 1).alignment = align
            ws.cell(2, 1, "วัน/คาบ").border = border
            ws.cell(2, 1).alignment = align

            group_credits = sum(s["credit"] for s in self.subjects if s["group"] == group)
            ws.cell(1, 2, f"รวมหน่วยกิต: {group_credits}").alignment = align

            for p in range(1, 12):
                c = ws.cell(2, p + 1, str(p))
                c.border = border
                c.alignment = align

            for i, d in enumerate(days_en, 3):
                d_th = day_en_to_th[d]
                c = ws.cell(i, 1, d_th)
                c.border = border
                c.alignment = align

            df, _ = schedule_room(group, self.subjects, slots_used, teacher_slots, self.locks, self.priority_mapping)

            for i, d in enumerate(days_en, 3):
                for p in range(1, 12):
                    txt = df.at[d, p]
                    if isinstance(txt, dict):
                        room_num = convert_room_letter_to_number(txt['room'], self.priority_mapping)
                        teacher_name = txt['teacher'].split()[0]
                        display = f"รหัส:{txt['code']}\nครู:{teacher_name}\nห้อง:{room_num}"
                    else:
                        display = txt
                    cell = ws.cell(i, p + 1, display)
                    cell.border = border
                    cell.alignment = align

            for c in range(1, 13):
                ws.column_dimensions[get_column_letter(c)].width = 15
            for r in range(1, 8):
                ws.row_dimensions[r].height = 45

        wb.save(path)



if __name__ == "__main__":
    root = tk.Tk()
    SchedulerApp(root)
    root.mainloop()